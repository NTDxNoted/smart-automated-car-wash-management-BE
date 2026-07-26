from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

rows = [
    ["No", "Workflow Code", "Workflow Name", "Description", "Num of Testcases"],
    [1, "WF-001", "Đăng nhập Member", "Đăng nhập bằng SĐT và mật khẩu, xác thực thông tin, phát hành JWT và chuyển tới hệ thống khách hàng.", 8],
    [2, "WF-002", "Trang chủ khách hàng", "Hiển thị profile, hạng thành viên, điểm và các chức năng chính cho Member.", 6],
    [3, "WF-003", "Hồ sơ cá nhân", "Member xem và cập nhật thông tin cá nhân được phép.", 7],
    [4, "WF-004", "Đặt lịch và quản lý booking", "Gộp toàn bộ luồng: quản lý xe, chọn dịch vụ, đặt lịch, áp dụng ưu đãi/điểm, xác nhận hóa đơn, xem lịch sử booking và xem catalog reward trong booking.", 24],
    [5, "WF-005", "Chi tiết đơn đặt lịch", "Member xem chi tiết booking, theo dõi trạng thái và hủy booking khi còn đủ điều kiện.", 10],
    [6, "WF-006", "Thông báo", "Hiển thị thông báo booking, thanh toán và trạng thái xử lý cho Member/Admin.", 5],
    [7, "WF-007", "Đăng nhập Admin & Dashboard", "Admin đăng nhập, nhận JWT và vào dashboard quản trị mặc định.", 9],
    [8, "WF-008", "Quản lý khách hàng", "Admin xem danh sách khách hàng, tìm kiếm và mở chi tiết Member.", 9],
    [9, "WF-009", "Quản lý đặt lịch, xử lý và thanh toán", "Admin xem booking, xử lý check-in/đối soát biển số, cập nhật trạng thái và xác nhận thanh toán.", 16],
    [10, "WF-010", "Quản lý dịch vụ", "Admin CRUD dịch vụ và cấu hình giá/thời lượng/loại xe.", 12],
    [11, "WF-011", "Quản lý khuyến mãi", "Admin CRUD promotion, bật/tắt và theo dõi usage.", 12],
    [12, "WF-012", "Quản lý hạng thành viên", "Admin xem và cập nhật ngưỡng tier.", 7],
    [13, "WF-013", "Báo cáo tổng quan", "Tổng hợp doanh thu, số booking, KPI và các chỉ số báo cáo gồm RFM, tier distribution, occupancy và promotion ROI.", 12],
    [14, "WF-014", "Báo cáo Loyalty", "Tổng hợp điểm đã cộng, đổi, hết hạn và hoàn trả cho hệ thống.", 8],
]

wb = Workbook()
ws = wb.active
ws.title = 'Workflow Summary'

for row in rows:
    ws.append(row)

header_fill = PatternFill('solid', fgColor='D9EAF7')
header_font = Font(bold=True)
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row_idx in range(1, len(rows) + 1):
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

for col_idx in range(1, 6):
    ws.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx == 4 else 15

ws['A1'].value = 'AutoWash Pro - Workflow Catalogue'
ws.merge_cells('A1:E1')
ws['A1'].font = Font(size=13, bold=True)
ws['A1'].fill = PatternFill('solid', fgColor='E2EFDA')
ws['A1'].alignment = Alignment(horizontal='center')

ws2 = wb.create_sheet('Notes')
ws2['A1'] = 'Ghi chú'
ws2['A2'] = 'Num of Testcases là số bước kiểm thử/ thao tác QA cần thực hiện cho từng workflow.'
ws2['A3'] = 'File được tạo tự động từ catalogue workflow của dự án.'
for cell in ws2['A1:A3']:
    cell[0].font = Font(bold=True)

out_path = Path(__file__).with_name('Workflow_Catalogue.xlsx')
out_path = str(out_path)
wb.save(out_path)
print(out_path)
